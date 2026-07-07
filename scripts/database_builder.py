from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


class DatabaseBuilder:

    def __init__(self):
        self.output = Path("output")
        self.output.mkdir(exist_ok=True)

    def save(self, bets_df, selections_df):

        output_file = self.output / "CBAS_Database.xlsx"

        # Save Bets and Selections sheets
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

            bets_df.to_excel(
                writer,
                sheet_name="Bets",
                index=False
            )

            selections_df.to_excel(
                writer,
                sheet_name="Selections",
                index=False
            )

        # Open workbook again to add Summary sheet
        wb = load_workbook(output_file)

        if "Summary" in wb.sheetnames:
            del wb["Summary"]

        ws = wb.create_sheet("Summary")

        # Title
        ws["A1"] = "CHEZAGAME BETTING ANALYTICS SYSTEM"
        ws["A1"].font = Font(size=16, bold=True)

        # Statistics
        total_bets = len(bets_df)

        total_stake = pd.to_numeric(
            bets_df["Stake"],
            errors="coerce"
        ).fillna(0).sum() if "Stake" in bets_df.columns else 0

        total_return = pd.to_numeric(
            bets_df["Return"],
            errors="coerce"
        ).fillna(0).sum() if "Return" in bets_df.columns else 0

        wins = (
            (bets_df["Result"] == "Won").sum()
            if "Result" in bets_df.columns
            else 0
        )

        losses = (
            (bets_df["Result"] == "Lost").sum()
            if "Result" in bets_df.columns
            else 0
        )

        profit = total_return - total_stake

        roi = (
            (profit / total_stake) * 100
            if total_stake > 0
            else 0
        )

        summary = [
            ("Total Bets", total_bets),
            ("Winning Bets", wins),
            ("Losing Bets", losses),
            ("Total Stake", round(total_stake, 2)),
            ("Total Return", round(total_return, 2)),
            ("Profit/Loss", round(profit, 2)),
            ("ROI (%)", round(roi, 2))
        ]

        row = 3

        for label, value in summary:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        wb.save(output_file)

        print(f"\nDatabase saved to: {output_file}")